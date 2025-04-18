import re
import os
import secrets
import shutil
from shutil import copyfile
from pathlib import Path

import tkinter
import customtkinter as ctk
import openpyxl as xl

# ---------- customtkinter init -----------------

root = ctk.CTk()

# ---------- Variables and CONSTANTS -----------------

# directory variables
path_strvar = ctk.StringVar(value='')
path_to_xlsx = Path()
path_to_xlsx_mistakes = Path()
project_dir = Path(__file__).parent.resolve()

# memory_file

# memory_file_name = 'memory.txt'
path_to_memory_txt = project_dir / 'memory.txt'

# is there valid memory.txt in project_dir:
memory_file_boolvar = ctk.BooleanVar(value=False)

# working with {mistakes}.xlsx:
record_boolvar = ctk.BooleanVar(value=False)
mistakes_file_chosen_boolvar = ctk.BooleanVar(value=False)
# {mistakes_file}.xlsx name (if it will be recorded)
mistakes_xlsx_name_strvar = ctk.StringVar(value='')

# WorkBook variables
file_name_strvar = ctk.StringVar(value='')  # {file}.xlsx name (WorkBook)
sheet_name_strvar = ctk.StringVar(value='')  # WorkSheet name
start_intvar = ctk.IntVar(value=0)
end_intvar = ctk.IntVar(value=0)
step_intvar = ctk.IntVar(value=200)

file_xlsx_read = xl.workbook.workbook.Workbook()
sheet = file_xlsx_read.create_sheet('fake_sheet')
files_xlsx_list = []
sheets_list = []  # list of WorkShits of WorkBook
start_end_list = []  # list of values for START/END comboboxes
rows_mixed_list = []  # cleared mixed list of rows numbers for test
mistakes_list = []  # list of WorkBook's mistakes rows

# GUI variables

# variable for mouse events on frame_2 correct work
check_mode_boolvar = ctk.BooleanVar(value=True)

# counter variables
words_done_qty_intvar = ctk.IntVar(value=0)
words_right_intvar = ctk.IntVar(value=0)
words_wrong_intvar = ctk.IntVar(value=0)
words_qty_intvar = ctk.IntVar(value=0)
percent_intvar = ctk.IntVar(value=0)

# Colors
GREY = '#D7ccc8'
GREY_DARK = '#a1887f'
GREY_LIGHT = '#efebe9'
BLUE = '#90caf9'
RED = '#FFB6C1'
GREEN = '#aed581'
FONT_LIGHT = 'white'
FONT_DARK = 'black'
FRAME_BG = '#eceff1'
YELLOW_BTN = '#FFE0B2'
YELLOW_BTN_HOVER = '#ffcc80'
FONT_BROWN = '#5d4037'
BROWN_BTN = '#bcaaa4'
BROWN_BTN_HOVER = '#a1887f'
SWITCH_BTN = '#6d4c41'


# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
#                                   *           *           *   &
#   --------- Functions ---------      *     *     *     *      &
#                                         *           *         &
# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&


def start_program():
    memory_file_boolvar_analyse()
    if memory_file_boolvar.get():
        memory_setting()
    # If memory_file missed, default stays: ''
    frame_1.anime()


def memory_file_boolvar_analyse():
    """
    Search for memory_file in cwd.
    Change memory_file_boolvar respectively.
    """
    REQUIRED_LINES = 6
    non_empty_lines = 0

    if not path_to_memory_txt.exists():
        return

    path_strvar = None

    try:
        with open(path_to_memory_txt) as file:
            for line in file:
                if line.strip():
                    non_empty_lines += 1
                if line.startswith("file_name_strvar="):
                    file_name_strvar = line.split(
                        "file_name_strvar=")[1].strip()
                if line.startswith("sheet_name_strvar="):
                    sheet_name_strvar = line.split(
                        "sheet_name_strvar=")[1].strip()
                if line.startswith("start_intvar="):
                    start_intvar = line.split("start_intvar=")[1].strip()
                if line.startswith("end_intvar="):
                    end_intvar = line.split("end_intvar=")[1].strip()
                if line.startswith("path_strvar="):
                    path_strvar = line.split("path_strvar=")[1].strip()
                if line.startswith("step_intvar="):
                    step_intvar = line.split("step_intvar=")[1].strip()
    except Exception as e:
        print(f"memory.txt read error: {e}")

    path_to_xlsx = Path(path_strvar) / file_name_strvar

    if not (
        non_empty_lines == REQUIRED_LINES and
        path_to_xlsx.exists() and
        start_intvar.isdigit() and
        end_intvar.isdigit() and
        step_intvar.isdigit()
    ):
        return

    workbook = xl.load_workbook(path_to_xlsx)

    if sheet_name_strvar not in workbook.sheetnames:
        return

    memory_file_boolvar.set(True)


def memory_setting():
    """
    Fill all start canvas's comboboxes at the beginning of mainloop.
    """
    global files_xlsx_list
    memory_data = memory_file_import()  # get dict from memory.txt
    # paths:
    path_strvar.set(value=memory_data['path_strvar'])
    assert_paths()
    # make list, that contains actually filenames.xlsx in paths:
    files_xlsx_list = create_files_xlsx_list()
    combo_file.configure(values=files_xlsx_list)
    # step:
    step_intvar.set(memory_data['step_intvar'])
    combo_step.set(str(step_intvar.get()))

    # Checking, if file.xlsx haven't been deleted:
    if memory_data['file_name_strvar'] in files_xlsx_list:
        # combo_file stage:
        file_name_strvar.set(value=memory_data['file_name_strvar'])
        combo_file.set(file_name_strvar.get())
        assert_file_xlsx_read()

        # combo_sheet stage:
        sheet_name_strvar.set(memory_data['sheet_name_strvar'])
        combo_sheet.set(sheet_name_strvar.get())
        combo_sheet.configure(values=file_xlsx_read.sheetnames)
        assert_sheet()

        # cobmo START/END var:
        start_end_combos_fill()

        start_intvar.set(memory_data['start_intvar'])
        combo_start.set(str(start_intvar.get()))
        end_intvar.set(memory_data['end_intvar'])
        combo_end.set(str(end_intvar.get()))

        words_qty_calc()
        btn_start_text()
        btn_start_check()


def assert_paths():
    global path_to_xlsx, path_to_xlsx_mistakes
    path_to_xlsx = Path(path_strvar.get())
    path_to_xlsx_mistakes = path_to_xlsx / 'mistakes'


def memory_file_import() -> dict:
    """
    Return dict with memory settings from the last using of the program.
    """
    with open(path_to_memory_txt) as file:
        # list of rows (local)
        text = file.readlines()
    data = {}
    for line in text:
        # List unpack from string, separated from '\n'
        # and split by '='; [k, val]:
        k, val = line.strip().split('=')
        if val.isnumeric():  # 'int' -> int. NB: float stays str
            val = int(val)
        data[k] = val
    return data


def create_files_xlsx_list() -> list:
    """
    Return list of all suitable files.xlsx in the user's path.
    """
    files_xlsx_list_local = []
    for file in path_to_xlsx.iterdir():
        match_xlsx_files = re.search('.xlsx', file.name)
        if match_xlsx_files:
            # filling list with files .xlsx names from path_to_xlsx
            files_xlsx_list_local.append(file.name)

    if path_to_xlsx_mistakes.exists():
        # Search for .xlsx files in Mistakes directory
        for file in path_to_xlsx_mistakes.iterdir():
            match_xlsx_files = re.search('.xlsx', file.name)
            if match_xlsx_files:
                # filling list with files .xlsx names from path_to_xlsx
                files_xlsx_list_local.append(file.name)
    else:
        path_to_xlsx_mistakes.mkdir(exist_ok=True)

    # Filter list from hidden and temporary files:
    files_xlsx_list_local = [
        _ for _ in files_xlsx_list_local if _[0] not in ('~', '.', '$')]

    return files_xlsx_list_local


def assert_file_xlsx_read():
    """
    Asserts global workbook object,
    changes mistakes_file_chosen_boolvar
    """
    global file_xlsx_read
    match = re.search('MSTK', file_name_strvar.get())
    if match:
        file_xlsx_read = xl.load_workbook(
            f"{str(path_to_xlsx_mistakes.absolute())}/{file_name_strvar.get()}"
        )
        mistakes_file_chosen_boolvar.set(True)
    else:
        file_xlsx_read = xl.load_workbook(
            f"{str(path_to_xlsx.absolute())}/{file_name_strvar.get()}")
        mistakes_file_chosen_boolvar.set(False)


def assert_sheet():
    global sheet
    sheet = file_xlsx_read[sheet_name_strvar.get()]


def combo_file_choice(choice: str):
    global sheets_list
    file_name_strvar.set(value=choice)
    assert_file_xlsx_read()
    sheets_list = file_xlsx_read.sheetnames
    combo_sheet.configure(values=sheets_list)
    combo_sheet.set(sheets_list[0])
    # assertion to different globals on the line of following comboboxes:
    combo_sheet_choice(sheets_list[0])


def combo_sheet_choice(choice):
    # global sheet
    sheet_name_strvar.set(choice)
    assert_sheet()
    start_end_combos_fill()


def start_end_combos_fill():
    """
    Create start_end_list.
    Pass it to START/END comboboxes.
    """
    global start_end_list
    # global files_xlsx_list, sheet
    start_end_list = []

    # Filling start_end_list:
    # NB: start_end_list is the list of str! (it's for START/END comboboxes),
    # At first, create local list of all rows in the sheet, that contain words:
    item = step = step_intvar.get()
    list_all_rows = []
    exceptions = (None, 'zzz')
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(column=1, row=row).value not in exceptions:
            list_all_rows.append(row)

    # If dictionary file is empty: user should fill comboboxes again,
    # beginning with combo_file:
    if len(list_all_rows) == 0:
        combos_clear()
        btn_start_deactivate()
        combo_file.configure(values=files_xlsx_list)
        combo_file.set("Your dictionary is empty!")

    else:
        # start value:
        start_end_list.append('2')
        # fill start_end_list with a step to maximum:
        while item < list_all_rows[-1]:
            start_end_list.append(str(item))
            item += step

        # add finish value of list, which often isn't multiple of step_intvar:
        # NB: if it's so, it will not duplicate, because of condition '<'
        start_end_list.append(str(list_all_rows[-1]))

        if len(start_end_list) == 1:
            # set the only value to values
            combo_start.configure(values=start_end_list)
            combo_end.configure(values=start_end_list)
            combo_start.set(start_end_list[0])
            combo_end.set(start_end_list[0])

        else:
            combo_start.configure(values=start_end_list[:-1])
            combo_end.configure(values=start_end_list[1:])
            # penultimate value of the list
            combo_start.set(start_end_list[-2])
            combo_end.set(start_end_list[-1])  # last value of the list

        start_intvar.set(int(combo_start.get()))
        end_intvar.set(int(combo_end.get()))
        words_qty_calc()
        btn_start_text()
        btn_start_check()


def words_qty_calc():
    if start_intvar.get() == end_intvar.get() == 2:
        words_qty_intvar.set(1)
    elif start_intvar.get() == end_intvar.get() != 2:
        words_qty_intvar.set(0)
    elif start_intvar.get() == 2 or end_intvar.get() % step_intvar.get() != 0:
        words_qty_intvar.set(
            int((end_intvar.get() - start_intvar.get()) / 2 + 1))
    else:
        words_qty_intvar.set(int((end_intvar.get() - start_intvar.get()) / 2))


def btn_start_check():
    if words_qty_intvar.get() <= 0:
        btn_start_deactivate()
    else:
        btn_start_activate()


def btn_start_activate():
    btn_start.configure(hover=True)
    btn_start.configure(command=start_exam)
    btn_start_text()
    btn_start.configure(text_color=FONT_BROWN)
    btn_start.focus()


def btn_start_deactivate():
    btn_start.configure(hover=False)
    btn_start.configure(command=None)
    btn_start.configure(text="Check Your Settings")
    btn_start.configure(text_color=GREY_DARK)


def clear_entry_path():
    path_strvar.set('')


def pass_entry_path():
    combos_clear()
    btn_start_deactivate()
    if not Path(entry_path.get()).exists() or entry_path.get() == '':
        combo_file.set("This path is not exist!")
    else:
        combo_file.set('Choose / Create a dictionary file')
        path_strvar.set(entry_path.get())
        assert_paths()
        combo_file.configure(values=create_files_xlsx_list())


def combos_clear():
    """
    CLear values and sets of all Workbook comboboxes
    """
    combo_file.configure(values='')
    combo_file.set('')
    combo_sheet.configure(values='')
    combo_sheet.set('')
    combo_start.configure(values='')
    combo_start.set('')
    combo_end.configure(values='')
    combo_end.set('')


def combo_start_choice(choice):
    start_intvar.set(int(choice))
    words_qty_calc()
    btn_start_text()
    btn_start_check()


def combo_end_choice(choice):
    end_intvar.set(int(choice))
    words_qty_calc()
    btn_start_text()
    btn_start_check()


def btn_start_text():
    if words_qty_intvar.get() == 1:
        btn_start.configure(text=f"START: {words_qty_intvar.get()} word")
    else:
        btn_start.configure(text=f"START: {words_qty_intvar.get()} words")


def create_template_xlsx():
    if entry_path.get() == '':
        combo_file.set("Fill the path first & press PASS")
    else:
        combos_clear()
        btn_start_deactivate()
        combo_file.set("'template.xlsx' has been created!")
        combo_file.configure(values=create_files_xlsx_list())
        copyfile('template.xlsx', Path(entry_path.get()) / 'template.xlsx')


def combo_step_choice(choice):
    step_intvar.set(int(choice))
    start_end_combos_fill()


def change_1_2():
    frame_1.place_forget()
    frame_2.anime()


def check():
    """
    Shows translation and transcription
    """
    label_translation.configure(text_color=FONT_DARK)
    label_translation.configure(
        text=str(sheet.cell(
            column=3,
            row=rows_mixed_list[words_done_qty_intvar.get()]).value).strip())
    label_transcription.configure(
        text=str(sheet.cell(
            column=2,
            row=rows_mixed_list[words_done_qty_intvar.get()]).value).strip())
    label_word.focus_set()  # focus for keyboard control
    check_mode_boolvar.set(False)


def record_xlsx():
    """
    Writing mistakes.xlsx
    """
    if not path_to_xlsx_mistakes.exists():
        path_to_xlsx_mistakes.mkdir()

    shutil.copyfile(Path(path_to_xlsx) / ('template.xlsx'),
                    Path(path_to_xlsx_mistakes) /
                    mistakes_xlsx_name_strvar.get())
    workbook_mistakes = xl.load_workbook(Path(path_to_xlsx_mistakes) /
                                         mistakes_xlsx_name_strvar.get())
    sheet_mistakes_book = workbook_mistakes.active

    # writing cells to sheet_mistakes_book:
    row_counter = 2
    for row in mistakes_list:
        sheet_mistakes_book.cell(row=row_counter, column=1).value = sheet.cell(
            row=row, column=1).value
        sheet_mistakes_book.cell(row=row_counter, column=2).value = sheet.cell(
            row=row, column=2).value
        sheet_mistakes_book.cell(row=row_counter, column=3).value = sheet.cell(
            row=row, column=3).value
        row_counter += 2

    # saving workbook_mistakes to path_to_xlsx_mistakes:
    workbook_mistakes.save(Path(path_to_xlsx_mistakes) /
                           mistakes_xlsx_name_strvar.get())


def change_2_win():
    label_result_win.configure(
        text=f"RESULT: {round(
            (words_right_intvar.get() / words_qty_intvar.get()) * 100)} %")
    frame_2.place_forget()
    frame_3_win.anime()
    btn_restart_win.focus_set()


def change_2_obvious_no_record():
    label_result_obvious.configure(
        text=f"RESULT: {round(
            (words_right_intvar.get() / words_qty_intvar.get()) * 100)} %")
    label_result_info_obvious.configure(
        text=f"The rest of words:"
        f"\n{words_wrong_intvar.get()} out of {words_qty_intvar.get()}")
    label_result_info_obvious.place(relx=0.5, rely=0.5, anchor='c')
    frame_2.place_forget()
    btn_restart_obvious.focus_set()
    frame_3_obvious.anime()


def change_2_obvious_record():
    record_xlsx()
    label_result_obvious.configure(
        text=f"RESULT: {round(
            (words_right_intvar.get() / words_qty_intvar.get()) * 100)} %")
    label_result_info_obvious.configure(
        text=f"The rest of words:"
        f"\n{words_wrong_intvar.get()} out of {words_qty_intvar.get()}")
    label_result_info_plus_obvious.configure(
        text=f"Saved: {mistakes_xlsx_name_strvar.get()}")
    label_result_info_obvious.place(relx=0.5, rely=0.4, anchor='c')
    label_result_info_plus_obvious.place(relx=0.5, rely=0.75, anchor='c')
    frame_2.place_forget()
    btn_restart_obvious.focus_set()
    frame_3_obvious.anime()


def analyse():
    finish_bool = words_done_qty_intvar.get() + 1 == len(rows_mixed_list)
    # check for finish win
    if finish_bool and not mistakes_list:
        change_2_win()
    # check for finish obvious without mistakes recording
    elif finish_bool and mistakes_list and not record_boolvar.get():
        change_2_obvious_no_record()
    # check for finish obvious with mistakes recording
    elif finish_bool and mistakes_list and record_boolvar.get():
        change_2_obvious_record()
    else:
        words_done_qty_intvar.set(words_done_qty_intvar.get() + 1)
        another_word()


def right():
    if check_mode_boolvar.get():
        check()
    else:
        words_right_intvar.set(words_right_intvar.get() + 1)
        label_know_number.configure(text=str(words_right_intvar.get()))
        check_mode_boolvar.set(True)
        analyse()


def wrong():

    # global mistakes_list
    if check_mode_boolvar.get():
        check()
    else:
        words_wrong_intvar.set(words_wrong_intvar.get() + 1)
        label_dont_know_number.configure(text=str(words_wrong_intvar.get()))
        # saving mistakes to list
        # (always happens, regardless record_boolvar.get())
        mistakes_list.append(rows_mixed_list[words_done_qty_intvar.get()])
        check_mode_boolvar.set(True)
        analyse()


def defaults():
    global sheets_list, start_end_list, rows_mixed_list
    global mistakes_list, files_xlsx_list

    # directory variables
    memory_file_boolvar.set(False)  # memory_file in cwd
    file_name_strvar.set(value='')  # file.xlsx name (WorkBook)
    sheet_name_strvar.set('')  # sheet name (WorkSheet)

    mistakes_xlsx_name_strvar.set('')  # mistakes file name

    # WorkBook variables
    files_xlsx_list = []
    sheets_list = []  # list of WorkBook' Sheet' names
    start_end_list = []  # list of START/END comboboxes values
    start_intvar.set(0)
    end_intvar.set(0)
    rows_mixed_list = []  # cleared mixed list of row's numbers for exam
    # list of WorkBook' rows, words from which haven't been passed.
    mistakes_list = []

    # counter variables
    words_done_qty_intvar.set(0)
    words_right_intvar.set(0)
    words_wrong_intvar.set(0)
    words_qty_intvar.set(0)
    percent_intvar.set(0)

    label_know_number.configure(text='0')
    label_dont_know_number.configure(text='0')
    record_boolvar.set(False)
    btn_switch.configure(text="MISTAKES\nRECORDING\n<-OFF->")
    btn_switch.configure(fg_color=GREY)
    label_result_info_plus_obvious.place_forget()


def change_3obvious_1():
    frame_3_obvious.place_forget()
    defaults()
    start_program()


def change_3win_1():
    frame_3_win.place_forget()
    defaults()
    start_program()


def memory_file_write():
    memory_data = dict()
    memory_data['file_name_strvar'] = file_name_strvar.get()
    memory_data['sheet_name_strvar'] = sheet_name_strvar.get()
    memory_data['start_intvar'] = start_intvar.get()
    memory_data['end_intvar'] = end_intvar.get()
    memory_data['path_strvar'] = path_strvar.get()
    memory_data['step_intvar'] = step_intvar.get()
    with open(path_to_memory_txt, 'w') as file:
        for k, val in memory_data.items():
            file.write(f"{k}={val}\n")


def switch_record_boolvar():
    if record_boolvar.get():
        record_boolvar.set(False)
        btn_switch.configure(text="MISTAKES\nRECORDING\n<-OFF->")
        btn_switch.configure(fg_color=GREY)
    else:
        record_boolvar.set(True)
        btn_switch.configure(text="MISTAKES\nRECORDING\n<-ON->")
        btn_switch.configure(fg_color=RED)


def another_word():
    label_word.configure(
        text=str(sheet.cell(
            column=1, row=rows_mixed_list
            [words_done_qty_intvar.get()]).value).strip())
    progress_bar.set(words_done_qty_intvar.get() / words_qty_intvar.get())
    percent_intvar.set(
        round((words_done_qty_intvar.get() / words_qty_intvar.get()) * 100))
    label_percents.configure(text=f"{percent_intvar.get()} %")
    label_words_qty.configure(
        text=f"{words_done_qty_intvar.get() + 1} out of "
        f"{words_qty_intvar.get()}")
    label_translation.configure(
        text='--->> check (click, DOWN, \'S\') <<---', text_color=GREY_DARK)
    label_transcription.configure(text='')
    label_translation.focus_set()  # focus for keyboard control


def start_exam():
    global rows_mixed_list
    # global sheet
    memory_file_write()
    #  creation of mistakes_xlsx_name_strvar:
    match = re.search('MSTK', file_name_strvar.get())  # finding mistakes files
    if match:
        mistakes_xlsx_name_strvar.set(
            f"{file_name_strvar.get()[:-5]}-"
            f"({start_intvar.get()}-{end_intvar.get()}).xlsx")
    else:
        mistakes_xlsx_name_strvar.set(
            f"MSTK_{file_name_strvar.get()[:-5]}-"
            f"({start_intvar.get()}-{end_intvar.get()}).xlsx")
    # list of selected range's rows:
    ordered_list = []
    start_loop = start_intvar.get()
    if mistakes_file_chosen_boolvar.get() or not_filled_range():
        # For mistakes files or not filled ranges: end is included
        for i in range(end_intvar.get() - start_loop + 1):
            ordered_list.append(start_loop)
            start_loop += 1
    else:
        # end value doesn't include: it will be first in next range
        # (otherwise first / last word will be duplicated in adjacent ranges)
        for i in range(end_intvar.get() - start_loop):
            ordered_list.append(start_loop)
            start_loop += 1
    #  formation of mixed and cleared list
    rows_mixed_list = []
    while len(ordered_list) != 0:
        v = secrets.choice(ordered_list)
        exceptions = [None]
        if sheet.cell(column=1, row=v).value not in exceptions:
            rows_mixed_list.append(v)
        ordered_list.remove(v)

    words_qty_intvar.set(len(rows_mixed_list))
    another_word()
    change_1_2()


def not_filled_range():
    if end_intvar.get() % step_intvar.get() != 0:
        return True
    else:
        return False


# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
#    *                                                     *    &
#       *  ----------------- GUI ---------------------  *       &
#    *                                                     *    &
# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

root.title("Vocabulary")

# window's appearance in the middle of the screen
root_width = int(0.4 * root.winfo_screenwidth())
root_height = int(0.5 * root.winfo_screenheight())
display_width = root.winfo_screenwidth()
display_height = root.winfo_screenheight()

left = int((display_width / 2) - (root_width / 2))
top = int((display_height / 2) - (root_height / 2))
root.geometry(f'{root_width}x{root_height}+{left}+{top}')

is_wayland = ("WAYLAND_DISPLAY" in os.environ or
              "HYPRLAND_INSTANCE_SIGNATURE" in os.environ)

if is_wayland:
    try:
        root.attributes('-type', 'dialog')
    # Specifically catch TclError which is raised by Tkinter
    except tkinter.TclError as e:
        print(f"Warning: Could not set window attributes: {e}")
    root.resizable(False, False)
    root.after(100, lambda: root.wm_attributes('-type', 'utility'))

# most popular scope size (others sizes will be got from it by a coefficient)
scope_base = int(round(root_height ** .43 + 2))
# wraplength parameter for labels
wraplength = int(round(root_width * .95))


# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#      Class Slide Panel (customization of CTkFrame class)
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@


class SlidePanel(ctk.CTkFrame):
    """
    Modernization of ctk frame class (animated appearance)
    """

    def __init__(self, direction_down=True):  # Class args to make class object
        # NB: all keys defaults should be exactly in super().__init__(..):
        super().__init__(master=root,
                         width=int(root_width),
                         height=int(root_height),
                         fg_color=FRAME_BG,
                         corner_radius=0)  # Args, inherited from parent class
        self.y_fly = None  # PyCharm advice
        self.slide = direction_down
        self.delta = 2
        if self.slide:
            self.y_start = - root_height - 2
            self.delta = self.delta
        else:
            self.y_start = root_height + 2

    def move_up(self):
        self.y_fly -= self.delta
        if self.y_fly > 0:
            self.place(x=0, y=self.y_fly)
            self.after(1, self.move_up)

    def move_down(self):
        self.y_fly += self.delta
        if self.y_fly < 0:
            self.place(x=0, y=self.y_fly)
            self.after(1, self.move_down)

    def anime(self):
        self.y_fly = self.y_start
        if self.slide:
            self.move_down()
        else:
            self.move_up()
        self.place(x=0, y=0)


# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#                 First canvas
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

frame_1 = SlidePanel()

# ----------- layout START Button -----------------

# This is not the highest widget positionally,
# but it's coded first to make it №1 in focus order.
btn_start = ctk.CTkButton(
    frame_1,
    font=('Calibri', scope_base * 2, 'bold'),
    text="Check Your Settings",
    text_color=GREY_DARK,
    corner_radius=0,
    fg_color=YELLOW_BTN,  # rest / active button's color
    hover_color=YELLOW_BTN_HOVER,  # hover button color
    command=None,
    hover=False
)

btn_start.bind('<space>', lambda e: start_exam())
btn_start.place(relx=.005, rely=.75, relheight=.245, relwidth=.99)

# ----------- layout File.xlsx ----------------------------------------------

canvas_file = ctk.CTkCanvas(frame_1, bg=GREY_LIGHT, )
canvas_file.place(relx=.005, rely=.005, relheight=.489, relwidth=.692)

# ----------- Frame_path  -----------------

frame_path = ctk.CTkFrame(
    canvas_file,
    fg_color=GREY,
)
frame_path.place(relx=.05, rely=.08, relwidth=.9, relheight=.24)

label_path = ctk.CTkLabel(
    frame_path, text="Path to your file's folder:",
    font=('Calibri', scope_base),
    fg_color=GREY,
    text_color=FONT_BROWN,
)
label_path.place(relx=.05, rely=.1, relwidth=.9, relheight=.4)

btn_clear_entry = ctk.CTkButton(
    frame_path,
    text='CLEAR',
    font=('Calibri', scope_base, 'bold'),
    text_color=FONT_BROWN,
    # rest / active button's color:
    fg_color=GREY_LIGHT,
    # hover button color:
    hover_color=RED,
    command=clear_entry_path,
)
btn_clear_entry.place(relx=.02, rely=.1, relwidth=.2, relheight=.4)

btn_path_pass = ctk.CTkButton(
    frame_path,
    text='PASS',
    font=('Calibri', scope_base, 'bold'),
    text_color=FONT_BROWN,
    fg_color=GREY_LIGHT,
    hover_color=RED,
    command=pass_entry_path,
)
btn_path_pass.place(relx=.78, rely=.1, relwidth=.2, relheight=.4)

entry_path = ctk.CTkEntry(
    frame_path,
    font=('Calibri', int(round(scope_base * .8))),
    border_color=GREY,
    justify='c',
    textvariable=path_strvar,
    # colors:
    fg_color="white",
    text_color="black",
)
entry_path.place(relx=.02, rely=.53, relwidth=.96, relheight=.4)

# ----------- Frame_file -----------------

frame_file = ctk.CTkFrame(
    canvas_file,
    fg_color=GREY,
)
frame_file.place(relx=.05, rely=.39, relwidth=.9, relheight=.24)

label_file = ctk.CTkLabel(
    frame_file,
    text="Choose file:",
    font=('Calibri', scope_base),
    fg_color=GREY,
    text_color=FONT_BROWN,
)
label_file.pack(pady=5)

combo_file = ctk.CTkComboBox(
    frame_file,
    font=('Calibri', scope_base),
    dropdown_font=('Calibri', scope_base),
    button_color=BROWN_BTN,
    button_hover_color=BROWN_BTN_HOVER,
    bg_color=GREY,
    border_color=GREY,
    justify='c',
    values=create_files_xlsx_list(),
    command=combo_file_choice,
    # colors:
    fg_color="white",
    text_color="black",
    dropdown_fg_color="white",
    dropdown_text_color="black",
    dropdown_hover_color="lightgray"
)

combo_file.set(file_name_strvar.get())
combo_file.pack(fill='both', padx=40)

# ----------- Frame_sheet -----------------

frame_sheet = ctk.CTkFrame(
    canvas_file,
    fg_color=GREY,
)
frame_sheet.place(relx=.05, rely=.7, relwidth=.9, relheight=.24)

label_sheet = ctk.CTkLabel(
    frame_sheet, text="Choose sheet:",
    font=('Calibri', scope_base),
    fg_color=GREY,
    text_color=FONT_BROWN,
)
label_sheet.pack(pady=5)

combo_sheet = ctk.CTkComboBox(
    frame_sheet,
    font=('Calibri', scope_base),
    dropdown_font=('Calibri', scope_base),
    button_color=BROWN_BTN,
    button_hover_color=BROWN_BTN_HOVER,
    bg_color=GREY,
    border_color=GREY,
    justify='c',
    values=sheets_list,
    command=combo_sheet_choice,
    # colors:
    fg_color="white",
    text_color="black",
    dropdown_fg_color="white",
    dropdown_text_color="black",
    dropdown_hover_color="lightgray"
)

combo_sheet.set(sheet_name_strvar.get())
combo_sheet.pack(fill='both', padx=40)

# ----------- layout Start-End ----------------------------------------------

canvas_start_end = ctk.CTkCanvas(frame_1, bg=GREY_LIGHT, )
canvas_start_end.place(relx=.005, rely=.5, relheight=.244, relwidth=.692)

# ----------- Frame_start -----------------

frame_start = ctk.CTkFrame(
    canvas_start_end,
    fg_color=GREY,
)
frame_start.place(relx=.05, rely=.25, relwidth=.42, relheight=.5)

label_start = ctk.CTkLabel(
    frame_start,
    text="Start:",
    font=('Calibri', scope_base),
    bg_color=GREY,
    text_color=FONT_BROWN,
    anchor='center',
    height=20,
)
label_start.pack(pady=5)

combo_start = ctk.CTkComboBox(
    frame_start,
    font=('Calibri', scope_base),
    dropdown_font=('Calibri', scope_base),
    button_color=BROWN_BTN,
    button_hover_color=BROWN_BTN_HOVER,
    bg_color=GREY,
    border_color=GREY,
    justify='c',
    command=combo_start_choice,
    # not to view default logo, when the file_name_strvar.get() == '':
    values=[],
    # colors:
    fg_color="white",
    text_color="black",
    dropdown_fg_color="white",
    dropdown_text_color="black",
    dropdown_hover_color="lightgray"
)

combo_start.set('')  # empty field, when the file_name_strvar.get() == ''

combo_start.pack(fill='both', padx=10)

# ----------- Frame_end -----------------

frame_end = ctk.CTkFrame(
    canvas_start_end,
    fg_color=GREY,
)
frame_end.place(relx=.53, rely=.25, relwidth=.42, relheight=.48)

label_end = ctk.CTkLabel(
    frame_end, text="End:",
    font=('Calibri', scope_base), bg_color=GREY,
    text_color=FONT_BROWN, anchor='center', height=20,
)
label_end.pack(pady=5)

combo_end = ctk.CTkComboBox(
    frame_end,
    font=('Calibri', scope_base),
    dropdown_font=('Calibri', scope_base),
    button_color=BROWN_BTN,
    button_hover_color=BROWN_BTN_HOVER,
    bg_color=GREY,
    border_color=GREY,
    justify='c',
    command=combo_end_choice,
    # not to view default logo, when the file_name_strvar.get() == '':
    values=[],
    # colors:
    fg_color="white",
    text_color="black",
    dropdown_fg_color="white",
    dropdown_text_color="black",
    dropdown_hover_color="lightgray"
)

combo_end.set('')  # empty field, when the file_name_strvar.get() == ''

combo_end.pack(fill='both', padx=10)

# ----------- layout Create template.xlsx ---------------------------------

canvas_template = ctk.CTkCanvas(frame_1, bg=GREY_LIGHT, )
canvas_template.place(relx=.7, rely=.005, relheight=.192, relwidth=0.295)

# ----------- button create template.xlsx -----------------

btn_template = ctk.CTkButton(
    canvas_template, text='CREATE NEW\ntemplate.xlsx',
    font=('Calibri', scope_base, 'bold'),
    text_color=FONT_BROWN,
    fg_color=GREY,  # rest / active button's color
    hover_color=RED,  # hover button color
    command=create_template_xlsx,
)

btn_template.place(relx=.1, rely=.2, relwidth=.8, relheight=.6)

# ----------- layout Switch ----------------------------------------------

canvas_switch = ctk.CTkCanvas(frame_1, bg=GREY_LIGHT, )
canvas_switch.place(relx=.7, rely=.203, relheight=.291, relwidth=0.295)

# ----------- Frame Record Switch -----------------

btn_switch = ctk.CTkButton(
    canvas_switch,
    text="MISTAKES\nRECORDING\n<-OFF->",
    font=('Calibri', int(round(scope_base * 1.1)),
          'bold'),
    text_color=FONT_BROWN,
    fg_color=GREY,  # rest / active button's color
    hover_color=RED,  # hover button color
    command=switch_record_boolvar,
)
btn_switch.place(relx=.1, rely=.1, relwidth=.8, relheight=.8)

# ----------- layout Step ----------------------------------------------

canvas_step = ctk.CTkCanvas(frame_1, bg=GREY_LIGHT, )
canvas_step.place(relx=.7, rely=.5, relheight=.244, relwidth=0.295)

# ----------- Frame Step choice -----------------

frame_step = ctk.CTkFrame(
    canvas_step,
    fg_color=GREY,
)
frame_step.place(relx=.1, rely=.25, relwidth=.8, relheight=.5)

label_step = ctk.CTkLabel(
    frame_step, text="STEP:",
    font=('Calibri', scope_base), bg_color=GREY,
    text_color=FONT_BROWN, anchor='center', height=20,
)
label_step.pack(pady=5)

combo_step = ctk.CTkComboBox(
    frame_step,
    font=('Calibri', scope_base),
    dropdown_font=('Calibri', scope_base),
    button_color=BROWN_BTN,
    button_hover_color=BROWN_BTN_HOVER,
    bg_color=GREY,
    border_color=GREY,
    justify='c',
    command=combo_step_choice,
    values=['20', '60', '100', '200', '300',
            '400', '600', '800', '1000', '2000'],
    # colors:
    fg_color="white",
    text_color="black",
    dropdown_fg_color="white",
    dropdown_text_color="black",
    dropdown_hover_color="lightgray"
)

combo_step.set('200')

combo_step.pack(fill='both', padx=10)

# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#                 Second canvas
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

frame_2 = SlidePanel(direction_down=False)
# frame_2.place(x=0, y=0)  # dev tool

# ----------- row 1 -----------------

label_words_qty = ctk.CTkLabel(
    frame_2,
    font=('Calibri', int(round(scope_base * 1.2))),
    bg_color=GREY,
    anchor='center',
    text_color=FONT_DARK,
)
label_words_qty.place(relx=.005, rely=.005, relwidth=.658, relheight=.07)

label_percents = ctk.CTkLabel(
    frame_2, bg_color=GREY,
    font=('Calibri', int(round(scope_base * 1.2))),
    text_color=FONT_DARK, anchor='center',
)
label_percents.place(relx=.668, rely=.005, relwidth=.326, relheight=.07)

# ----------- row 2 -----------------

label_know = ctk.CTkLabel(
    frame_2, text="Know (LMK, LEFT, 'A')",
    font=('Calibri', scope_base),
    bg_color=GREY_DARK,
    text_color=FONT_LIGHT, anchor='center',
)
label_know.place(relx=.005, rely=.082, relwidth=.492, relheight=.06)

label_dont_know = ctk.CTkLabel(
    frame_2, text="Don't (RMK, RIGHT, 'D')",
    font=('Calibri', scope_base),
    bg_color=GREY_DARK,
    text_color=FONT_LIGHT,
    anchor='center',
)
label_dont_know.place(relx=.503, rely=.082, relwidth=.492, relheight=.06)

# ----------- row 3 -----------------

label_know_number = ctk.CTkLabel(
    frame_2, text=str(words_right_intvar.get()),
    font=('Calibri',
          int(round(scope_base * 1.2))),
    bg_color=GREY,
    text_color=FONT_DARK, anchor='center',
)
label_know_number.place(relx=.005, rely=.149, relwidth=.492, relheight=.07)

label_dont_know_number = ctk.CTkLabel(
    frame_2,
    text=str(words_wrong_intvar.get()),
    font=('Calibri', int(
        round(scope_base * 1.2))),
    bg_color=GREY,
    text_color=FONT_DARK, anchor='center',
)
label_dont_know_number.place(
    relx=.503, rely=.149, relwidth=.492, relheight=.07)

# ----------- row 4 -----------------

label_word = ctk.CTkLabel(frame_2,
                          font=('Calibri', int(round(scope_base * 1.8))),
                          bg_color=BLUE, text_color=FONT_DARK,
                          anchor='center', wraplength=wraplength)

label_word.place(relx=.005, rely=.226, relwidth=.99, relheight=.2)

label_word.bind("<Button-1>", lambda e: right())
label_word.bind("<Left>", lambda e: right())
label_word.bind("<a>", lambda e: right())
label_word.bind("<A>", lambda e: right())

label_word.bind("<Button-3>", lambda e: wrong())
label_word.bind("<Right>", lambda e: wrong())
label_word.bind("<d>", lambda e: wrong())
label_word.bind("<D>", lambda e: wrong())

# ----------- row 5 -----------------

label_translation = ctk.CTkLabel(
    frame_2,
    font=('Calibri',
          int(round(scope_base * 1.6))),
    bg_color=GREY,
    text_color=GREY_DARK,
    anchor='center',
    wraplength=wraplength
)

label_translation.place(relx=.005, rely=.433, relwidth=.99, relheight=.31)

# control buttons for translation
label_translation.bind("<Button-1>", lambda e: check())
label_translation.bind("<Button-3>", lambda e: check())
label_translation.bind("<Down>", lambda e: check())
label_translation.bind("<s>", lambda e: check())
label_translation.bind("<S>", lambda e: check())

# ----------- row 6 -----------------

label_transcription = ctk.CTkLabel(frame_2,
                                   font=('Calibri',
                                         int(round(scope_base * 1.6))),
                                   bg_color=GREY_DARK,
                                   text_color=FONT_LIGHT, anchor='center',
                                   wraplength=wraplength)

label_transcription.place(relx=.005, rely=.75, relwidth=.99, relheight=.2)

# ----------- row 7 -----------------

progress_bar = ctk.CTkProgressBar(frame_2, bg_color=FRAME_BG,
                                  corner_radius=0,
                                  fg_color='white', progress_color=GREY)

progress_bar.place(relx=.0, rely=.96, relheight=.035, relwidth=1)

# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#                 Finish Canvas Obvious
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

frame_3_obvious = SlidePanel()
# frame_3_obvious.place(x=0, y=0)  # dev tool

# ----------- layout Top Bar -----------------

label_result_obvious = ctk.CTkLabel(
    frame_3_obvious,
    font=('Calibri', int(
        round(scope_base * 1.33)), 'bold'),
    anchor='center',
    text_color=FONT_LIGHT,
    bg_color=GREY_DARK,
)
label_result_obvious.place(relx=.005, rely=.005, relheight=.175, relwidth=.99)

# ----------- layout Middle Bar -----------------

canvas_middle_bar_obvious = ctk.CTkCanvas(frame_3_obvious, bg=GREY_LIGHT, )
canvas_middle_bar_obvious.place(
    relx=.005, rely=.188, relheight=.555, relwidth=.99)

label_result_info_obvious = ctk.CTkLabel(
    canvas_middle_bar_obvious,
    font=(
        'Calibri', scope_base * 2, 'bold'),
    anchor='center', text_color=FONT_BROWN,
    bg_color=GREY_LIGHT
)

label_result_info_plus_obvious = ctk.CTkLabel(
    canvas_middle_bar_obvious,
    font=('Calibri', int(
        round(scope_base * 1.2)),),
    anchor='center', text_color=FONT_BROWN,
    bg_color=GREY_LIGHT, wraplength=wraplength
)

# ----------- layout ReSTART Button -----------------

btn_restart_obvious = ctk.CTkButton(
    frame_3_obvious, text='ReSTART',
    font=('Calibri', scope_base * 2, 'bold'),
    corner_radius=0,
    text_color=FONT_BROWN,
    fg_color=YELLOW_BTN,  # rest / active button's color
    hover_color=YELLOW_BTN_HOVER,  # hover button color
    command=change_3obvious_1
)

btn_restart_obvious.bind('<space>', lambda e: change_3obvious_1())
btn_restart_obvious.place(relx=.005, rely=.75, relheight=.245, relwidth=.99)

# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#                 Finish Canvas Win
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

frame_3_win = SlidePanel()
# frame_3_win.place(x=0, y=0)  # dev tool

# ----------- layout Top Bar -----------------

label_result_win = ctk.CTkLabel(
    frame_3_win, font=('Calibri', int(round(scope_base * 1.33)), 'bold'),
    anchor='center', text_color=FONT_BROWN, width=392, height=60,
    bg_color=GREY_LIGHT
)
label_result_win.place(relx=.005, rely=.005, relheight=.175, relwidth=.99)

# ----------- layout Middle Bar -----------------

canvas_middle_bar_win = ctk.CTkCanvas(frame_3_win, bg=GREEN, height=288, )
canvas_middle_bar_win.place(relx=.005, rely=.188, relheight=.555, relwidth=.99)

label_result_info_win = ctk.CTkLabel(
    canvas_middle_bar_win, font=('Calibri', scope_base * 4, 'bold'),
    text='Success!', anchor='center', text_color=FONT_LIGHT,
)

label_result_info_win.place(relx=0.5, rely=0.5, anchor='c')

# ----------- layout ReSTART Button -----------------

btn_restart_win = ctk.CTkButton(
    frame_3_win, text='ReSTART', font=('Calibri', scope_base * 2, 'bold'),
    corner_radius=0,
    height=80, text_color=FONT_BROWN,
    fg_color=YELLOW_BTN,  # rest / active button's color
    hover_color=YELLOW_BTN_HOVER,  # hover button color
    command=change_3win_1
)
btn_restart_win.bind('<space>', lambda e: change_3win_1())
btn_restart_win.place(relx=.005, rely=.75, relheight=.245, relwidth=.99)

# ---------- Start Program -----------------

if __name__ == '__main__':
    start_program()
    root.mainloop()
