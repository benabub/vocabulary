import os
import platform
import re
import secrets
import shutil
from pathlib import Path
from shutil import copyfile
import customtkinter as ctk
import openpyxl as xl


class VocabularyApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Vocabulary")

        # Определение окружения
        self.is_wayland = self.detect_wayland()
        self.is_macos = platform.system() == 'Darwin'
        self.is_windows = platform.system() == 'Windows'

        # Настройка окна
        self.setup_window()

        # Инициализация переменных
        self.init_variables()

        # Создание интерфейса
        self.create_widgets()

        # Запуск программы
        self.start_program()

    def detect_wayland(self):
        """Определяем Wayland окружение"""
        wayland_vars = [
            os.environ.get("XDG_SESSION_TYPE") == "wayland",
            "WAYLAND_DISPLAY" in os.environ,
            "HYPRLAND_INSTANCE_SIGNATURE" in os.environ
        ]
        return any(wayland_vars)

    def setup_window(self):
        """Настройка окна в зависимости от ОС"""
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        if self.is_wayland:
            # Borderless режим для Wayland/Hyprland
            self.overrideredirect(True)
            self.geometry(f"{screen_width}x{screen_height}+0+0")
            self.current_width = screen_width
            self.current_height = screen_height
        else:
            # Стандартное поведение для других ОС
            width = int(0.4 * screen_width)
            height = int(0.5 * screen_height)
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            self.geometry(f"{width}x{height}+{x}+{y}")
            self.current_width = width
            self.current_height = height

        # Расчет производных параметров
        self.scope_base = int(round(self.current_height ** 0.43 + 2))
        self.wraplength = int(round(self.current_width * 0.95))

        # Привязка обработчика изменения размера
        self.bind("<Configure>", self.on_window_resize)

    def init_variables(self):
        """Инициализация всех переменных"""
        # directory variables
        self.path_strvar = ctk.StringVar(value='')
        self.path_to_xlsx = Path()
        self.path_to_xlsx_mistakes = Path()

        # memory_file
        self.memory_file_name = 'memory.txt'
        self.memory_file_boolvar = ctk.BooleanVar(value=False)

        # working with {mistakes}.xlsx:
        self.record_boolvar = ctk.BooleanVar(value=False)
        self.mistakes_file_chosen_boolvar = ctk.BooleanVar(value=False)
        self.mistakes_xlsx_name_strvar = ctk.StringVar(value='')

        # WorkBook variables
        self.file_name_strvar = ctk.StringVar(value='')
        self.sheet_name_strvar = ctk.StringVar(value='')
        self.start_intvar = ctk.IntVar(value=0)
        self.end_intvar = ctk.IntVar(value=0)
        self.step_intvar = ctk.IntVar(value=200)

        self.file_xlsx_read = xl.workbook.workbook.Workbook()
        self.sheet = self.file_xlsx_read.create_sheet('fake_sheet')
        self.files_xlsx_list = []
        self.sheets_list = []
        self.start_end_list = []
        self.rows_mixed_list = []
        self.mistakes_list = []

        # GUI variables
        self.check_mode_boolvar = ctk.BooleanVar(value=True)

        # counter variables
        self.words_done_qty_intvar = ctk.IntVar(value=0)
        self.words_right_intvar = ctk.IntVar(value=0)
        self.words_wrong_intvar = ctk.IntVar(value=0)
        self.words_qty_intvar = ctk.IntVar(value=0)
        self.percent_intvar = ctk.IntVar(value=0)

        # Colors
        self.GREY = '#D7ccc8'
        self.GREY_DARK = '#a1887f'
        self.GREY_LIGHT = '#efebe9'
        self.BLUE = '#90caf9'
        self.RED = '#FFB6C1'
        self.GREEN = '#aed581'
        self.FONT_LIGHT = 'white'
        self.FONT_DARK = 'black'
        self.FRAME_BG = '#eceff1'
        self.YELLOW_BTN = '#FFE0B2'
        self.YELLOW_BTN_HOVER = '#ffcc80'
        self.FONT_BROWN = '#5d4037'
        self.BROWN_BTN = '#bcaaa4'
        self.BROWN_BTN_HOVER = '#a1887f'
        self.SWITCH_BTN = '#6d4c41'

    def on_window_resize(self, event):
        """Обработчик изменения размера окна"""
        if event.widget == self:
            self.current_width = self.winfo_width()
            self.current_height = self.winfo_height()
            self.scope_base = int(round(self.current_height ** 0.43 + 2))
            self.wraplength = int(round(self.current_width * 0.95))
            self.update_widgets()

    def update_widgets(self):
        """Обновление виджетов при изменении размеров"""
        if hasattr(self, 'main_label'):
            self.main_label.configure(wraplength=self.wraplength)
        if hasattr(self, 'label_translation'):
            self.label_translation.configure(wraplength=self.wraplength)
        if hasattr(self, 'label_transcription'):
            self.label_transcription.configure(wraplength=self.wraplength)

    def start_program(self):
        self.memory_file_boolvar_analyse()
        if self.memory_file_boolvar.get():
            self.memory_setting()
        self.frame_1.anime()

    def memory_file_boolvar_analyse(self):
        """Search for memory_file in cwd"""
        if Path(self.memory_file_name).exists():
            self.memory_file_boolvar.set(True)

    def memory_setting(self):
        """Fill all start canvas's comboboxes"""
        memory_data = self.memory_file_import()
        self.path_strvar.set(value=memory_data['path_strvar'])
        self.assert_paths()
        self.files_xlsx_list = self.create_files_xlsx_list()
        self.combo_file.configure(values=self.files_xlsx_list)
        self.step_intvar.set(memory_data['step_intvar'])
        self.combo_step.set(str(self.step_intvar.get()))

        if memory_data['file_name_strvar'] in self.files_xlsx_list:
            self.file_name_strvar.set(value=memory_data['file_name_strvar'])
            self.combo_file.set(self.file_name_strvar.get())
            self.assert_file_xlsx_read()

            self.sheet_name_strvar.set(memory_data['sheet_name_strvar'])
            self.combo_sheet.set(self.sheet_name_strvar.get())
            self.combo_sheet.configure(values=self.file_xlsx_read.sheetnames)
            self.assert_sheet()

            self.start_end_combos_fill()

            self.start_intvar.set(memory_data['start_intvar'])
            self.combo_start.set(str(self.start_intvar.get()))
            self.end_intvar.set(memory_data['end_intvar'])
            self.combo_end.set(str(self.end_intvar.get()))

            self.words_qty_calc()
            self.btn_start_text()
            self.btn_start_check()

    def memory_file_import(self):
        """Return dict with memory settings"""
        with open(self.memory_file_name) as file:
            text = file.readlines()
        data = {}
        for line in text:
            k, val = line.strip().split('=')
            if val.isnumeric():
                val = int(val)
            data[k] = val
        return data

    def assert_paths(self):
        self.path_to_xlsx = Path(self.path_strvar.get())
        self.path_to_xlsx_mistakes = self.path_to_xlsx / 'mistakes'

    def create_files_xlsx_list(self):
        """Return list of all suitable files.xlsx"""
        files_xlsx_list_local = []
        for file in self.path_to_xlsx.iterdir():
            if re.search('.xlsx', file.name):
                files_xlsx_list_local.append(file.name)

        for file in self.path_to_xlsx_mistakes.iterdir():
            if re.search('.xlsx', file.name):
                files_xlsx_list_local.append(file.name)

        return [f for f in files_xlsx_list_local if f[0] not in ('~', '.', '$')]

    def assert_file_xlsx_read(self):
        match = re.search('MSTK', self.file_name_strvar.get())
        if match:
            self.file_xlsx_read = xl.load_workbook(
                f"{str(self.path_to_xlsx_mistakes.absolute())}/{self.file_name_strvar.get()}")
            self.mistakes_file_chosen_boolvar.set(True)
        else:
            self.file_xlsx_read = xl.load_workbook(
                f"{str(self.path_to_xlsx.absolute())}/{self.file_name_strvar.get()}")
            self.mistakes_file_chosen_boolvar.set(False)

    def assert_sheet(self):
        self.sheet = self.file_xlsx_read[self.sheet_name_strvar.get()]

    def combo_file_choice(self, choice):
        self.file_name_strvar.set(value=choice)
        self.assert_file_xlsx_read()
        self.sheets_list = self.file_xlsx_read.sheetnames
        self.combo_sheet.configure(values=self.sheets_list)
        self.combo_sheet.set(self.sheets_list[0])
        self.combo_sheet_choice(self.sheets_list[0])

    def combo_sheet_choice(self, choice):
        self.sheet_name_strvar.set(choice)
        self.assert_sheet()
        self.start_end_combos_fill()

    def start_end_combos_fill(self):
        self.start_end_list = []
        list_all_rows = []
        exceptions = (None, 'zzz')

        for row in range(1, self.sheet.max_row + 1):
            if self.sheet.cell(column=1, row=row).value not in exceptions:
                list_all_rows.append(row)

        if len(list_all_rows) == 0:
            self.combos_clear()
            self.btn_start_deactivate()
            self.combo_file.configure(values=self.files_xlsx_list)
            self.combo_file.set("Your dictionary is empty!")
        else:
            self.start_end_list.append('2')
            item = step = self.step_intvar.get()

            while item < list_all_rows[-1]:
                self.start_end_list.append(str(item))
                item += step

            self.start_end_list.append(str(list_all_rows[-1]))

            if len(self.start_end_list) == 1:
                self.combo_start.configure(values=self.start_end_list)
                self.combo_end.configure(values=self.start_end_list)
                self.combo_start.set(self.start_end_list[0])
                self.combo_end.set(self.start_end_list[0])
            else:
                self.combo_start.configure(values=self.start_end_list[:-1])
                self.combo_end.configure(values=self.start_end_list[1:])
                self.combo_start.set(self.start_end_list[-2])
                self.combo_end.set(self.start_end_list[-1])

            self.start_intvar.set(int(self.combo_start.get()))
            self.end_intvar.set(int(self.combo_end.get()))
            self.words_qty_calc()
            self.btn_start_text()
            self.btn_start_check()

    def words_qty_calc(self):
        if self.start_intvar.get() == self.end_intvar.get() == 2:
            self.words_qty_intvar.set(1)
        elif self.start_intvar.get() == self.end_intvar.get() != 2:
            self.words_qty_intvar.set(0)
        elif self.start_intvar.get() == 2 or self.end_intvar.get() % self.step_intvar.get() != 0:
            self.words_qty_intvar.set(
                int((self.end_intvar.get() - self.start_intvar.get()) / 2 + 1))
        else:
            self.words_qty_intvar.set(
                int((self.end_intvar.get() - self.start_intvar.get()) / 2))

    def btn_start_check(self):
        if self.words_qty_intvar.get() <= 0:
            self.btn_start_deactivate()
        else:
            self.btn_start_activate()

    def btn_start_activate(self):
        self.btn_start.configure(hover=True)
        self.btn_start.configure(command=self.start_exam)
        self.btn_start_text()
        self.btn_start.configure(text_color=self.FONT_BROWN)
        self.btn_start.focus()

    def btn_start_deactivate(self):
        self.btn_start.configure(hover=False)
        self.btn_start.configure(command=None)
        self.btn_start.configure(text="Check Your Settings")
        self.btn_start.configure(text_color=self.GREY_DARK)

    def clear_entry_path(self):
        self.path_strvar.set('')

    def pass_entry_path(self):
        self.combos_clear()
        self.btn_start_deactivate()
        if not Path(self.entry_path.get()).exists() or self.entry_path.get() == '':
            self.combo_file.set("This path is not exist!")
        else:
            self.combo_file.set('Choose / Create a dictionary file')
            self.path_strvar.set(self.entry_path.get())
            self.assert_paths()
            self.combo_file.configure(values=self.create_files_xlsx_list())

    def combos_clear(self):
        self.combo_file.configure(values='')
        self.combo_file.set('')
        self.combo_sheet.configure(values='')
        self.combo_sheet.set('')
        self.combo_start.configure(values='')
        self.combo_start.set('')
        self.combo_end.configure(values='')
        self.combo_end.set('')

    def combo_start_choice(self, choice):
        self.start_intvar.set(int(choice))
        self.words_qty_calc()
        self.btn_start_text()
        self.btn_start_check()

    def combo_end_choice(self, choice):
        self.end_intvar.set(int(choice))
        self.words_qty_calc()
        self.btn_start_text()
        self.btn_start_check()

    def btn_start_text(self):
        if self.words_qty_intvar.get() == 1:
            self.btn_start.configure(
                text=f"START: {self.words_qty_intvar.get()} word")
        else:
            self.btn_start.configure(
                text=f"START: {self.words_qty_intvar.get()} words")

    def create_template_xlsx(self):
        if self.entry_path.get() == '':
            self.combo_file.set("Fill the path first & press PASS")
        else:
            self.combos_clear()
            self.btn_start_deactivate()
            self.combo_file.set("'template.xlsx' has been created!")
            self.combo_file.configure(values=self.create_files_xlsx_list())
            copyfile('template.xlsx', Path(
                self.entry_path.get()) / 'template.xlsx')

    def combo_step_choice(self, choice):
        self.step_intvar.set(int(choice))
        self.start_end_combos_fill()

    def change_1_2(self):
        self.frame_1.place_forget()
        self.frame_2.anime()

    def check(self):
        self.label_translation.configure(text_color=self.FONT_DARK)
        self.label_translation.configure(
            text=str(self.sheet.cell(column=3, row=self.rows_mixed_list[self.words_done_qty_intvar.get()]).value).strip())
        self.label_transcription.configure(
            text=str(self.sheet.cell(column=2, row=self.rows_mixed_list[self.words_done_qty_intvar.get()]).value).strip())
        self.label_word.focus_set()
        self.check_mode_boolvar.set(False)

    def record_xlsx(self):
        if not self.path_to_xlsx_mistakes.exists():
            self.path_to_xlsx_mistakes.mkdir()

        shutil.copyfile(Path('template.xlsx'), Path(self.path_to_xlsx_mistakes) /
                        self.mistakes_xlsx_name_strvar.get())
        workbook_mistakes = xl.load_workbook(Path(self.path_to_xlsx_mistakes) /
                                             self.mistakes_xlsx_name_strvar.get())
        sheet_mistakes_book = workbook_mistakes.active

        row_counter = 2
        for row in self.mistakes_list:
            sheet_mistakes_book.cell(row=row_counter, column=1).value = self.sheet.cell(
                row=row, column=1).value
            sheet_mistakes_book.cell(row=row_counter, column=2).value = self.sheet.cell(
                row=row, column=2).value
            sheet_mistakes_book.cell(row=row_counter, column=3).value = self.sheet.cell(
                row=row, column=3).value
            row_counter += 2

        workbook_mistakes.save(Path(self.path_to_xlsx_mistakes) /
                               self.mistakes_xlsx_name_strvar.get())

    def change_2_win(self):
        self.label_result_win.configure(
            text=f"RESULT: {round((self.words_right_intvar.get() / self.words_qty_intvar.get()) * 100)} %")
        self.frame_2.place_forget()
        self.frame_3_win.anime()
        self.btn_restart_win.focus_set()

    def change_2_obvious_no_record(self):
        self.label_result_obvious.configure(
            text=f"RESULT: {round((self.words_right_intvar.get() / self.words_qty_intvar.get()) * 100)} %")
        self.label_result_info_obvious.configure(
            text=f"The rest of words:\n{self.words_wrong_intvar.get()} out of {self.words_qty_intvar.get()}")
        self.label_result_info_obvious.place(relx=0.5, rely=0.5, anchor='c')
        self.frame_2.place_forget()
        self.btn_restart_obvious.focus_set()
        self.frame_3_obvious.anime()

    def change_2_obvious_record(self):
        self.record_xlsx()
        self.label_result_obvious.configure(
            text=f"RESULT: {round((self.words_right_intvar.get() / self.words_qty_intvar.get()) * 100)} %")
        self.label_result_info_obvious.configure(
            text=f"The rest of words:\n{self.words_wrong_intvar.get()} out of {self.words_qty_intvar.get()}")
        self.label_result_info_plus_obvious.configure(
            text=f"Saved: {self.mistakes_xlsx_name_strvar.get()}")
        self.label_result_info_obvious.place(relx=0.5, rely=0.4, anchor='c')
        self.label_result_info_plus_obvious.place(
            relx=0.5, rely=0.75, anchor='c')
        self.frame_2.place_forget()
        self.btn_restart_obvious.focus_set()
        self.frame_3_obvious.anime()

    def analyse(self):
        finish_bool = self.words_done_qty_intvar.get() + 1 == len(self.rows_mixed_list)

        if finish_bool and not self.mistakes_list:
            self.change_2_win()
        elif finish_bool and self.mistakes_list and not self.record_boolvar.get():
            self.change_2_obvious_no_record()
        elif finish_bool and self.mistakes_list and self.record_boolvar.get():
            self.change_2_obvious_record()
        else:
            self.words_done_qty_intvar.set(
                self.words_done_qty_intvar.get() + 1)
            self.another_word()

    def right(self):
        if self.check_mode_boolvar.get():
            self.check()
        else:
            self.words_right_intvar.set(self.words_right_intvar.get() + 1)
            self.label_know_number.configure(
                text=str(self.words_right_intvar.get()))
            self.check_mode_boolvar.set(True)
            self.analyse()

    def wrong(self):
        if self.check_mode_boolvar.get():
            self.check()
        else:
            self.words_wrong_intvar.set(self.words_wrong_intvar.get() + 1)
            self.label_dont_know_number.configure(
                text=str(self.words_wrong_intvar.get()))
            self.mistakes_list.append(
                self.rows_mixed_list[self.words_done_qty_intvar.get()])
            self.check_mode_boolvar.set(True)
            self.analyse()

    def defaults(self):
        self.memory_file_boolvar.set(False)
        self.file_name_strvar.set(value='')
        self.sheet_name_strvar.set('')
        self.mistakes_xlsx_name_strvar.set('')
        self.files_xlsx_list = []
        self.sheets_list = []
        self.start_end_list = []
        self.start_intvar.set(0)
        self.end_intvar.set(0)
        self.rows_mixed_list = []
        self.mistakes_list = []
        self.words_done_qty_intvar.set(0)
        self.words_right_intvar.set(0)
        self.words_wrong_intvar.set(0)
        self.words_qty_intvar.set(0)
        self.percent_intvar.set(0)
        self.label_know_number.configure(text='0')
        self.label_dont_know_number.configure(text='0')
        self.record_boolvar.set(False)
        self.btn_switch.configure(text="MISTAKES\nRECORDING\n<-OFF->")
        self.btn_switch.configure(fg_color=self.GREY)
        self.label_result_info_plus_obvious.place_forget()

    def change_3obvious_1(self):
        self.frame_3_obvious.place_forget()
        self.defaults()
        self.start_program()

    def change_3win_1(self):
        self.frame_3_win.place_forget()
        self.defaults()
        self.start_program()

    def memory_file_write(self):
        memory_data = dict()
        memory_data['file_name_strvar'] = self.file_name_strvar.get()
        memory_data['sheet_name_strvar'] = self.sheet_name_strvar.get()
        memory_data['start_intvar'] = self.start_intvar.get()
        memory_data['end_intvar'] = self.end_intvar.get()
        memory_data['path_strvar'] = self.path_strvar.get()
        memory_data['step_intvar'] = self.step_intvar.get()
        with open(self.memory_file_name, 'w') as file:
            for k, val in memory_data.items():
                file.write(f"{k}={val}\n")

    def switch_record_boolvar(self):
        if self.record_boolvar.get():
            self.record_boolvar.set(False)
            self.btn_switch.configure(text="MISTAKES\nRECORDING\n<-OFF->")
            self.btn_switch.configure(fg_color=self.GREY)
        else:
            self.record_boolvar.set(True)
            self.btn_switch.configure(text="MISTAKES\nRECORDING\n<-ON->")
            self.btn_switch.configure(fg_color=self.RED)

    def another_word(self):
        self.label_word.configure(
            text=str(self.sheet.cell(column=1, row=self.rows_mixed_list[self.words_done_qty_intvar.get()]).value).strip())
        self.progress_bar.set(
            self.words_done_qty_intvar.get() / self.words_qty_intvar.get())
        self.percent_intvar.set(
            round((self.words_done_qty_intvar.get() / self.words_qty_intvar.get()) * 100))
        self.label_percents.configure(text=f"{self.percent_intvar.get()} %")
        self.label_words_qty.configure(
            text=f"{self.words_done_qty_intvar.get() + 1} out of {self.words_qty_intvar.get()}")
        self.label_translation.configure(
            text='--->> check (click, DOWN, \'S\') <<---', text_color=self.GREY_DARK)
        self.label_transcription.configure(text='')
        self.label_translation.focus_set()

    def start_exam(self):
        self.memory_file_write()
        match = re.search('MSTK', self.file_name_strvar.get())
        if match:
            self.mistakes_xlsx_name_strvar.set(
                f"{self.file_name_strvar.get()[:-5]}-({self.start_intvar.get()}-{self.end_intvar.get()}).xlsx")
        else:
            self.mistakes_xlsx_name_strvar.set(
                f"MSTK_{self.file_name_strvar.get()[:-5]}-({self.start_intvar.get()}-{self.end_intvar.get()}).xlsx")

        ordered_list = []
        start_loop = self.start_intvar.get()

        if self.mistakes_file_chosen_boolvar.get() or self.not_filled_range():
            for i in range(self.end_intvar.get() - start_loop + 1):
                ordered_list.append(start_loop)
                start_loop += 1
        else:
            for i in range(self.end_intvar.get() - start_loop):
                ordered_list.append(start_loop)
                start_loop += 1

        self.rows_mixed_list = []
        while len(ordered_list) != 0:
            v = secrets.choice(ordered_list)
            if self.sheet.cell(column=1, row=v).value not in [None]:
                self.rows_mixed_list.append(v)
            ordered_list.remove(v)

        self.words_qty_intvar.set(len(self.rows_mixed_list))
        self.another_word()
        self.change_1_2()

    def not_filled_range(self):
        return self.end_intvar.get() % self.step_intvar.get() != 0

    def create_widgets(self):
        """Создание всех виджетов интерфейса"""
        # Frame 1
        self.frame_1 = SlidePanel(self)

        # Start Button
        self.btn_start = ctk.CTkButton(
            self.frame_1,
            font=('Calibri', self.scope_base * 2, 'bold'),
            text="Check Your Settings",
            text_color=self.GREY_DARK,
            corner_radius=0,
            fg_color=self.YELLOW_BTN,
            hover_color=self.YELLOW_BTN_HOVER,
            command=None,
            hover=False
        )
        self.btn_start.bind('<space>', lambda e: self.start_exam())
        self.btn_start.place(relx=.005, rely=.75, relheight=.245, relwidth=.99)

        # File.xlsx Canvas
        self.canvas_file = ctk.CTkCanvas(self.frame_1, bg=self.GREY_LIGHT)
        self.canvas_file.place(relx=.005, rely=.005,
                               relheight=.489, relwidth=.692)

        # Frame Path
        self.frame_path = ctk.CTkFrame(
            self.canvas_file,
            fg_color=self.GREY,
        )
        self.frame_path.place(relx=.05, rely=.08, relwidth=.9, relheight=.24)

        self.label_path = ctk.CTkLabel(
            self.frame_path,
            text="Path to your file's folder:",
            font=('Calibri', self.scope_base),
            fg_color=self.GREY,
            text_color=self.FONT_BROWN,
        )
        self.label_path.place(relx=.05, rely=.1, relwidth=.9, relheight=.4)

        self.btn_clear_entry = ctk.CTkButton(
            self.frame_path,
            text='CLEAR',
            font=('Calibri', self.scope_base, 'bold'),
            text_color=self.FONT_BROWN,
            fg_color=self.GREY_LIGHT,
            hover_color=self.RED,
            command=self.clear_entry_path,
        )
        self.btn_clear_entry.place(
            relx=.02, rely=.1, relwidth=.2, relheight=.4)

        self.btn_path_pass = ctk.CTkButton(
            self.frame_path,
            text='PASS',
            font=('Calibri', self.scope_base, 'bold'),
            text_color=self.FONT_BROWN,
            fg_color=self.GREY_LIGHT,
            hover_color=self.RED,
            command=self.pass_entry_path,
        )
        self.btn_path_pass.place(relx=.78, rely=.1, relwidth=.2, relheight=.4)

        self.entry_path = ctk.CTkEntry(
            self.frame_path,
            font=('Calibri', int(round(self.scope_base * .8))),
            border_color=self.GREY,
            justify='c',
            textvariable=self.path_strvar
        )
        self.entry_path.place(relx=.02, rely=.53, relwidth=.96, relheight=.4)

        # Frame File
        self.frame_file = ctk.CTkFrame(
            self.canvas_file,
            fg_color=self.GREY,
        )
        self.frame_file.place(relx=.05, rely=.39, relwidth=.9, relheight=.24)

        self.label_file = ctk.CTkLabel(
            self.frame_file,
            text="Choose file:",
            font=('Calibri', self.scope_base),
            fg_color=self.GREY,
            text_color=self.FONT_BROWN,
        )
        self.label_file.pack(pady=5)

        self.combo_file = ctk.CTkComboBox(
            self.frame_file,
            font=('Calibri', self.scope_base),
            dropdown_font=('Calibri', self.scope_base),
            button_color=self.BROWN_BTN,
            button_hover_color=self.BROWN_BTN_HOVER,
            bg_color=self.GREY,
            border_color=self.GREY,
            justify='c',
            command=self.combo_file_choice,
        )
        self.combo_file.set(self.file_name_strvar.get())
        self.combo_file.pack(fill='both', padx=40)

        # Frame Sheet
        self.frame_sheet = ctk.CTkFrame(
            self.canvas_file,
            fg_color=self.GREY,
        )
        self.frame_sheet.place(relx=.05, rely=.7, relwidth=.9, relheight=.24)

        self.label_sheet = ctk.CTkLabel(
            self.frame_sheet,
            text="Choose sheet:",
            font=('Calibri', self.scope_base),
            fg_color=self.GREY,
            text_color=self.FONT_BROWN,
        )
        self.label_sheet.pack(pady=5)

        self.combo_sheet = ctk.CTkComboBox(
            self.frame_sheet,
            font=('Calibri', self.scope_base),
            dropdown_font=('Calibri', self.scope_base),
            button_color=self.BROWN_BTN,
            button_hover_color=self.BROWN_BTN_HOVER,
            bg_color=self.GREY,
            border_color=self.GREY,
            justify='c',
            command=self.combo_sheet_choice
        )
        self.combo_sheet.set(self.sheet_name_strvar.get())
        self.combo_sheet.pack(fill='both', padx=40)

        # Start-End Canvas
        self.canvas_start_end = ctk.CTkCanvas(self.frame_1, bg=self.GREY_LIGHT)
        self.canvas_start_end.place(
            relx=.005, rely=.5, relheight=.244, relwidth=.692)

        # Frame Start
        self.frame_start = ctk.CTkFrame(
            self.canvas_start_end,
            fg_color=self.GREY,
        )
        self.frame_start.place(relx=.05, rely=.25, relwidth=.42, relheight=.5)

        self.label_start = ctk.CTkLabel(
            self.frame_start,
            text="Start:",
            font=('Calibri', self.scope_base),
            bg_color=self.GREY,
            text_color=self.FONT_BROWN,
            anchor='center',
            height=20,
        )
        self.label_start.pack(pady=5)

        self.combo_start = ctk.CTkComboBox(
            self.frame_start,
            font=('Calibri', self.scope_base),
            dropdown_font=('Calibri', self.scope_base),
            button_color=self.BROWN_BTN,
            button_hover_color=self.BROWN_BTN_HOVER,
            bg_color=self.GREY,
            border_color=self.GREY,
            justify='c',
            command=self.combo_start_choice,
            values=[]
        )
        self.combo_start.set('')
        self.combo_start.pack(fill='both', padx=10)

        # Frame End
        self.frame_end = ctk.CTkFrame(
            self.canvas_start_end,
            fg_color=self.GREY,
        )
        self.frame_end.place(relx=.53, rely=.25, relwidth=.42, relheight=.48)

        self.label_end = ctk.CTkLabel(
            self.frame_end,
            text="End:",
            font=('Calibri', self.scope_base),
            bg_color=self.GREY,
            text_color=self.FONT_BROWN,
            anchor='center',
            height=20,
        )
        self.label_end.pack(pady=5)

        self.combo_end = ctk.CTkComboBox(
            self.frame_end,
            font=('Calibri', self.scope_base),
            dropdown_font=('Calibri', self.scope_base),
            button_color=self.BROWN_BTN,
            button_hover_color=self.BROWN_BTN_HOVER,
            bg_color=self.GREY,
            border_color=self.GREY,
            justify='c',
            command=self.combo_end_choice,
            values=[]
        )
        self.combo_end.set('')
        self.combo_end.pack(fill='both', padx=10)

        # Template Canvas
        self.canvas_template = ctk.CTkCanvas(self.frame_1, bg=self.GREY_LIGHT)
        self.canvas_template.place(
            relx=.7, rely=.005, relheight=.192, relwidth=0.295)

        # Create template button
        self.btn_template = ctk.CTkButton(
            self.canvas_template,
            text='CREATE NEW\ntemplate.xlsx',
            font=('Calibri', self.scope_base, 'bold'),
            text_color=self.FONT_BROWN,
            fg_color=self.GREY,
            hover_color=self.RED,
            command=self.create_template_xlsx,
        )
        self.btn_template.place(relx=.1, rely=.2, relwidth=.8, relheight=.6)

        # Switch Canvas
        self.canvas_switch = ctk.CTkCanvas(self.frame_1, bg=self.GREY_LIGHT)
        self.canvas_switch.place(
            relx=.7, rely=.203, relheight=.291, relwidth=0.295)

        # Record Switch
        self.btn_switch = ctk.CTkButton(
            self.canvas_switch,
            text="MISTAKES\nRECORDING\n<-OFF->",
            font=('Calibri', int(round(self.scope_base * 1.1)), 'bold'),
            text_color=self.FONT_BROWN,
            fg_color=self.GREY,
            hover_color=self.RED,
            command=self.switch_record_boolvar,
        )
        self.btn_switch.place(relx=.1, rely=.1, relwidth=.8, relheight=.8)

        # Step Canvas
        self.canvas_step = ctk.CTkCanvas(self.frame_1, bg=self.GREY_LIGHT)
        self.canvas_step.place(
            relx=.7, rely=.5, relheight=.244, relwidth=0.295)

        # Frame Step
        self.frame_step = ctk.CTkFrame(
            self.canvas_step,
            fg_color=self.GREY,
        )
        self.frame_step.place(relx=.1, rely=.25, relwidth=.8, relheight=.5)

        self.label_step = ctk.CTkLabel(
            self.frame_step,
            text="STEP:",
            font=('Calibri', self.scope_base),
            bg_color=self.GREY,
            text_color=self.FONT_BROWN,
            anchor='center',
            height=20,
        )
        self.label_step.pack(pady=5)

        self.combo_step = ctk.CTkComboBox(
            self.frame_step,
            font=('Calibri', self.scope_base),
            dropdown_font=('Calibri', self.scope_base),
            button_color=self.BROWN_BTN,
            button_hover_color=self.BROWN_BTN_HOVER,
            bg_color=self.GREY,
            border_color=self.GREY,
            justify='c',
            command=self.combo_step_choice,
            values=['20', '60', '100', '200', '300',
                    '400', '600', '800', '1000', '2000']
        )
        self.combo_step.set('200')
        self.combo_step.pack(fill='both', padx=10)

        # Frame 2
        self.frame_2 = SlidePanel(self, direction_down=False)

        # Row 1
        self.label_words_qty = ctk.CTkLabel(
            self.frame_2,
            font=('Calibri', int(round(self.scope_base * 1.2))),
            bg_color=self.GREY,
            anchor='center',
            text_color=self.FONT_DARK,
        )
        self.label_words_qty.place(
            relx=.005, rely=.005, relwidth=.658, relheight=.07)

        self.label_percents = ctk.CTkLabel(
            self.frame_2,
            bg_color=self.GREY,
            font=('Calibri', int(round(self.scope_base * 1.2))),
            text_color=self.FONT_DARK,
            anchor='center',
        )
        self.label_percents.place(
            relx=.668, rely=.005, relwidth=.326, relheight=.07)

        # Row 2
        self.label_know = ctk.CTkLabel(
            self.frame_2,
            text="Know (LMK, LEFT, 'A')",
            font=('Calibri', self.scope_base),
            bg_color=self.GREY_DARK,
            text_color=self.FONT_LIGHT,
            anchor='center',
        )
        self.label_know.place(relx=.005, rely=.082,
                              relwidth=.492, relheight=.06)

        self.label_dont_know = ctk.CTkLabel(
            self.frame_2,
            text="Don't (RMK, RIGHT, 'D')",
            font=('Calibri', self.scope_base),
            bg_color=self.GREY_DARK,
            text_color=self.FONT_LIGHT,
            anchor='center',
        )
        self.label_dont_know.place(
            relx=.503, rely=.082, relwidth=.492, relheight=.06)

        # Row 3
        self.label_know_number = ctk.CTkLabel(
            self.frame_2,
            text=str(self.words_right_intvar.get()),
            font=('Calibri', int(round(self.scope_base * 1.2))),
            bg_color=self.GREY,
            text_color=self.FONT_DARK,
            anchor='center',
        )
        self.label_know_number.place(
            relx=.005, rely=.149, relwidth=.492, relheight=.07)

        self.label_dont_know_number = ctk.CTkLabel(
            self.frame_2,
            text=str(self.words_wrong_intvar.get()),
            font=('Calibri', int(round(self.scope_base * 1.2))),
            bg_color=self.GREY,
            text_color=self.FONT_DARK,
            anchor='center',
        )
        self.label_dont_know_number.place(
            relx=.503, rely=.149, relwidth=.492, relheight=.07)

        # Row 4
        self.label_word = ctk.CTkLabel(
            self.frame_2,
            font=('Calibri', int(round(self.scope_base * 1.8))),
            bg_color=self.BLUE,
            text_color=self.FONT_DARK,
            anchor='center',
            wraplength=self.wraplength
        )
        self.label_word.place(relx=.005, rely=.226, relwidth=.99, relheight=.2)

        self.label_word.bind("<Button-1>", lambda e: self.right())
        self.label_word.bind("<Left>", lambda e: self.right())
        self.label_word.bind("<a>", lambda e: self.right())
        self.label_word.bind("<A>", lambda e: self.right())

        self.label_word.bind("<Button-3>", lambda e: self.wrong())
        self.label_word.bind("<Right>", lambda e: self.wrong())
        self.label_word.bind("<d>", lambda e: self.wrong())
        self.label_word.bind("<D>", lambda e: self.wrong())

        # Row 5
        self.label_translation = ctk.CTkLabel(
            self.frame_2,
            font=('Calibri', int(round(self.scope_base * 1.6))),
            bg_color=self.GREY,
            text_color=self.GREY_DARK,
            anchor='center',
            wraplength=self.wraplength
        )
        self.label_translation.place(
            relx=.005, rely=.433, relwidth=.99, relheight=.31)

        self.label_translation.bind("<Button-1>", lambda e: self.check())
        self.label_translation.bind("<Button-3>", lambda e: self.check())
        self.label_translation.bind("<Down>", lambda e: self.check())
        self.label_translation.bind("<s>", lambda e: self.check())
        self.label_translation.bind("<S>", lambda e: self.check())

        # Row 6
        self.label_transcription = ctk.CTkLabel(
            self.frame_2,
            font=('Calibri', int(round(self.scope_base * 1.6))),
            bg_color=self.GREY_DARK,
            text_color=self.FONT_LIGHT,
            anchor='center',
            wraplength=self.wraplength
        )
        self.label_transcription.place(
            relx=.005, rely=.75, relwidth=.99, relheight=.2)

        # Row 7
        self.progress_bar = ctk.CTkProgressBar(
            self.frame_2,
            bg_color=self.FRAME_BG,
            corner_radius=0,
            fg_color='white',
            progress_color=self.GREY
        )
        self.progress_bar.place(relx=.0, rely=.96, relheight=.035, relwidth=1)

        # Frame 3 Obvious
        self.frame_3_obvious = SlidePanel(self)

        # Top Bar
        self.label_result_obvious = ctk.CTkLabel(
            self.frame_3_obvious,
            font=('Calibri', int(round(self.scope_base * 1.33)), 'bold'),
            anchor='center',
            text_color=self.FONT_LIGHT,
            bg_color=self.GREY_DARK,
        )
        self.label_result_obvious.place(
            relx=.005, rely=.005, relheight=.175, relwidth=.99)

        # Middle Bar
        self.canvas_middle_bar_obvious = ctk.CTkCanvas(
            self.frame_3_obvious, bg=self.GREY_LIGHT)
        self.canvas_middle_bar_obvious.place(
            relx=.005, rely=.188, relheight=.555, relwidth=.99)

        self.label_result_info_obvious = ctk.CTkLabel(
            self.canvas_middle_bar_obvious,
            font=('Calibri', self.scope_base * 2, 'bold'),
            anchor='center',
            text_color=self.FONT_BROWN,
            bg_color=self.GREY_LIGHT
        )

        self.label_result_info_plus_obvious = ctk.CTkLabel(
            self.canvas_middle_bar_obvious,
            font=('Calibri', int(round(self.scope_base * 1.2)),),
            anchor='center',
            text_color=self.FONT_BROWN,
            bg_color=self.GREY_LIGHT,
            wraplength=self.wraplength
        )

        # ReSTART Button
        self.btn_restart_obvious = ctk.CTkButton(
            self.frame_3_obvious,
            text='ReSTART',
            font=('Calibri', self.scope_base * 2, 'bold'),
            corner_radius=0,
            text_color=self.FONT_BROWN,
            fg_color=self.YELLOW_BTN,
            hover_color=self.YELLOW_BTN_HOVER,
            command=self.change_3obvious_1
        )
        self.btn_restart_obvious.bind(
            '<space>', lambda e: self.change_3obvious_1())
        self.btn_restart_obvious.place(
            relx=.005, rely=.75, relheight=.245, relwidth=.99)

        # Frame 3 Win
        self.frame_3_win = SlidePanel(self)

        # Top Bar
        self.label_result_win = ctk.CTkLabel(
            self.frame_3_win,
            font=('Calibri', int(round(self.scope_base * 1.33)), 'bold'),
            anchor='center',
            text_color=self.FONT_BROWN,
            width=392,
            height=60,
            bg_color=self.GREY_LIGHT
        )
        self.label_result_win.place(
            relx=.005, rely=.005, relheight=.175, relwidth=.99)

        # Middle Bar
        self.canvas_middle_bar_win = ctk.CTkCanvas(
            self.frame_3_win, bg=self.GREEN, height=288)
        self.canvas_middle_bar_win.place(
            relx=.005, rely=.188, relheight=.555, relwidth=.99)

        self.label_result_info_win = ctk.CTkLabel(
            self.canvas_middle_bar_win,
            font=('Calibri', self.scope_base * 4, 'bold'),
            text='Success!',
            anchor='center',
            text_color=self.FONT_LIGHT,
        )
        self.label_result_info_win.place(relx=0.5, rely=0.5, anchor='c')

        # ReSTART Button
        self.btn_restart_win = ctk.CTkButton(
            self.frame_3_win,
            text='ReSTART',
            font=('Calibri', self.scope_base * 2, 'bold'),
            corner_radius=0,
            height=80,
            text_color=self.FONT_BROWN,
            fg_color=self.YELLOW_BTN,
            hover_color=self.YELLOW_BTN_HOVER,
            command=self.change_3win_1
        )
        self.btn_restart_win.bind('<space>', lambda e: self.change_3win_1())
        self.btn_restart_win.place(
            relx=.005, rely=.75, relheight=.245, relwidth=.99)


class SlidePanel(ctk.CTkFrame):
    """Modernization of ctk frame class (animated appearance)"""

    def __init__(self, master, direction_down=True):
        super().__init__(
            master=master,
            width=master.current_width,
            height=master.current_height,
            fg_color=master.FRAME_BG,
            corner_radius=0
        )
        self.y_fly = None
        self.slide = direction_down
        self.delta = 2
        if self.slide:
            self.y_start = - master.current_height - 2
        else:
            self.y_start = master.current_height + 2

    def move_up(self):
        self.y_fly -= self.delta
        if self.y_fly > 0:
            self.place(x=0, y=self.y_fly)
            self.after(1, self.move_up)

    def move_down(self):
        self.y_fly += self.delta
        if self.y_fly < 0:
            self.place(x=0, y=self.y_fly)
            self.after(1, self.move_down)

    def anime(self):
        self.y_fly = self.y_start
        if self.slide:
            self.move_down()
        else:
            self.move_up()
        self.place(x=0, y=0)


if __name__ == '__main__':
    app = VocabularyApp()
    app.mainloop()
